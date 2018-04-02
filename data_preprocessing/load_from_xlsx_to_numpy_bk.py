# load data from xlsx to npz
# put annotation_format.xlsx in data_collection and this file outside

import openpyxl as px
import numpy as np 

fps = 30
data_dir = 'data_collection/' 
xls = px.load_workbook(data_dir + 'annotation_format.xlsx')
sheet = xls.get_sheet_by_name(name='Sheet1')

target_dict = {'r':0, 'g':1, 'b':2, 'y':3, 'o':4, 'p':0}
num = 1
name_list = []
target_list = []
start_list = []
end_list = []

for row in sheet.iter_rows():
	if sheet.cell(column=5, row=num).value == 'N':
		num += 1
		continue
	name_list.append(sheet.cell(column=1, row=num).value)
	print(sheet.cell(column=2, row=num).value)
	target_list.append(target_dict[sheet.cell(column=2, row=num).value])
	# print(sheet.cell(column=3, row=num).value.minute)
	# print(sheet.cell(column=4, row=num).value.minute)
	# print(sheet.cell(column=3, row=num).value)
	# print(sheet.cell(column=4, row=num).value)
	# start = sheet.cell(column=3, row=num).value.hour * 60 + sheet.cell(column=3, row=num).value.minute
	# end = sheet.cell(column=4, row=num).value.hour * 60 + sheet.cell(column=4, row=num).value.minute
	if type(sheet.cell(column=3, row=num).value) == int:
		start = sheet.cell(column=3, row=num).value
		end = sheet.cell(column=4, row=num).value
	else:
		start = sheet.cell(column=3, row=num).value.minute * fps
		end = sheet.cell(column=4, row=num).value.minute * fps + fps
	start_list.append(start)
	end_list.append(end)
	num += 1

print(name_list)
print(target_list)
print(start_list)
print(end_list)
name_list = np.array(name_list)
target_list = np.array(target_list)
start_list = np.array(start_list)
end_list = np.array(end_list)
np.savez(data_dir + 'annotation.npz', name_list, target_list, start_list, end_list)
print('save annotation to npy sucessfully!')
