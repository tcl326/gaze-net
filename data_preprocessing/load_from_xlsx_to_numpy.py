# load data from xlsx to npz
# put annotation_format.xlsx in data_collection and this file outside

import openpyxl as px
import numpy as np

fps = 30
data_dir = ''
xls = px.load_workbook(data_dir + 'annotation_format_new.xlsx')


target_dict = {'w':0, 'g':1, 'b':2, 'y':3, 'o':4}
num = 1
name_list = []
target_list = []
start_list = []
end_list = []
hand_list = []

set = ['set1', 'set2', 'set3', 'set4']
for s in set:
	sheet = xls.get_sheet_by_name(name=s)
	num = 1
	print(s)
	for row in sheet.iter_rows():
		if num == 1:
			num += 1
			continue
		if sheet.cell(column=6, row=num).value == 'N':
			num += 1
			continue
		if sheet.cell(column=1, row=num).value == None:
			break
		tmp = s + '/' + sheet.cell(column=1, row=num).value.split('/')[-1]
		name_list.append(tmp)
		print(tmp)
		# print(sheet.cell(column=2, row=num).value)
		target_list.append(target_dict[sheet.cell(column=2, row=num).value])
		start = sheet.cell(column=3, row=num).value
		end = sheet.cell(column=4, row=num).value
		hand = sheet.cell(column=5, row=num).value
		start_list.append(start)
		end_list.append(end)
		hand_list.append(hand)
		num += 1

# print(name_list)
# print(target_list)
# print(start_list)
# print(end_list)
name_list = np.array(name_list)
target_list = np.array(target_list)
start_list = np.array(start_list)
end_list = np.array(end_list)
hand_list = np.array(hand_list)
print(len(end_list))
print(len(hand_list))
np.savez(data_dir + 'annotation.npz', name_list, target_list, start_list, end_list, hand_list)
print('save annotation to npy sucessfully!')
